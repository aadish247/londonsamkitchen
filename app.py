from flask import Flask, render_template, request, jsonify, flash, send_file
from flask_sqlalchemy import SQLAlchemy
from datetime import datetime, timedelta
import pandas as pd
import os
import calendar
import io

app = Flask(__name__)

# Use DATABASE_URL for production, fallback to SQLite for local development
database_url = os.environ.get('DATABASE_URL')
if database_url:
    # Handle Railway/Render database URL format
    if database_url.startswith("postgres://"):
        database_url = database_url.replace("postgres://", "postgresql://", 1)
    app.config['SQLALCHEMY_DATABASE_URI'] = database_url
else:
    app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///food_truck.db'

# Use environment variable for secret key
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'your-secret-key-here')
db = SQLAlchemy(app)


class Investment(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    investor_name = db.Column(db.String(100), nullable=False)
    amount = db.Column(db.Float, nullable=False)
    date = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)


class Expense(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    description = db.Column(db.String(200), nullable=False)
    amount = db.Column(db.Float, nullable=False)
    date = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)
    category = db.Column(db.String(100))


class Sale(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    amount = db.Column(db.Float, nullable=False)
    date = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)
    description = db.Column(db.String(200))


with app.app_context():
    db.create_all()


@app.route('/')
def index():
    total_investment = sum(inv.amount for inv in Investment.query.all())
    total_expenses = sum(exp.amount for exp in Expense.query.all())
    total_sales = sum(sale.amount for sale in Sale.query.all())

    investment_count = Investment.query.count()
    expense_count = Expense.query.count()
    sale_count = Sale.query.count()

    return render_template('index.html',
                           total_investment=total_investment,
                           total_expenses=total_expenses,
                           total_sales=total_sales,
                           investment_count=investment_count,
                           expense_count=expense_count,
                           sale_count=sale_count)


@app.route('/investments')
def investments():
    investments = Investment.query.order_by(Investment.date.desc()).all()
    total_investment = sum(inv.amount for inv in investments)
    return render_template('investments.html', investments=investments, total=total_investment)


@app.route('/add_investment', methods=['POST'])
def add_investment():
    data = request.json
    new_investment = Investment(
        investor_name=data['investor_name'],
        amount=float(data['amount']),
        date=datetime.strptime(data['date'], '%Y-%m-%d')
    )
    db.session.add(new_investment)
    db.session.commit()
    return jsonify({'status': 'success'})


@app.route('/expenses')
def expenses():
    expenses = Expense.query.order_by(Expense.date.desc()).all()
    total_expenses = sum(exp.amount for exp in expenses)
    return render_template('expenses.html', expenses=expenses, total=total_expenses)


@app.route('/add_expense', methods=['POST'])
def add_expense():
    data = request.json
    new_expense = Expense(
        description=data['description'],
        amount=float(data['amount']),
        category=data['category'],
        date=datetime.strptime(data['date'], '%Y-%m-%d')
    )
    db.session.add(new_expense)
    db.session.commit()
    return jsonify({'status': 'success'})


@app.route('/sales')
def sales():
    sales = Sale.query.order_by(Sale.date.desc()).all()
    total_sales = sum(sale.amount for sale in sales)
    return render_template('sales.html', sales=sales, total=total_sales)


@app.route('/add_sale', methods=['POST'])
def add_sale():
    data = request.json
    new_sale = Sale(
        amount=float(data['amount']),
        description=data['description'],
        date=datetime.strptime(data['date'], '%Y-%m-%d')
    )
    db.session.add(new_sale)
    db.session.commit()
    return jsonify({'status': 'success'})


@app.route('/delete_investment/<int:id>', methods=['POST'])
def delete_investment(id):
    investment = Investment.query.get_or_404(id)
    db.session.delete(investment)
    db.session.commit()
    return jsonify({'status': 'success'})


@app.route('/delete_expense/<int:id>', methods=['POST'])
def delete_expense(id):
    expense = Expense.query.get_or_404(id)
    db.session.delete(expense)
    db.session.commit()
    return jsonify({'status': 'success'})


@app.route('/delete_sale/<int:id>', methods=['POST'])
def delete_sale(id):
    sale = Sale.query.get_or_404(id)
    db.session.delete(sale)
    db.session.commit()
    return jsonify({'status': 'success'})


@app.route('/edit_investment/<int:id>', methods=['GET', 'POST'])
def edit_investment(id):
    investment = Investment.query.get_or_404(id)

    if request.method == 'POST':
        data = request.json
        investment.investor_name = data['investor_name']
        investment.amount = float(data['amount'])
        investment.date = datetime.strptime(data['date'], '%Y-%m-%d')
        db.session.commit()
        return jsonify({'status': 'success'})

    return jsonify({
        'id': investment.id,
        'investor_name': investment.investor_name,
        'amount': investment.amount,
        'date': investment.date.strftime('%Y-%m-%d')
    })


@app.route('/edit_expense/<int:id>', methods=['GET', 'POST'])
def edit_expense(id):
    expense = Expense.query.get_or_404(id)

    if request.method == 'POST':
        data = request.json
        expense.description = data['description']
        expense.amount = float(data['amount'])
        expense.category = data['category']
        expense.date = datetime.strptime(data['date'], '%Y-%m-%d')
        db.session.commit()
        return jsonify({'status': 'success'})

    return jsonify({
        'id': expense.id,
        'description': expense.description,
        'amount': expense.amount,
        'category': expense.category,
        'date': expense.date.strftime('%Y-%m-%d')
    })


@app.route('/edit_sale/<int:id>', methods=['GET', 'POST'])
def edit_sale(id):
    sale = Sale.query.get_or_404(id)

    if request.method == 'POST':
        data = request.json
        sale.description = data['description']
        sale.amount = float(data['amount'])
        sale.date = datetime.strptime(data['date'], '%Y-%m-%d')
        db.session.commit()
        return jsonify({'status': 'success'})

    return jsonify({
        'id': sale.id,
        'description': sale.description,
        'amount': sale.amount,
        'date': sale.date.strftime('%Y-%m-%d')
    })


@app.route('/export_data')
def export_data():
    try:
        # Create a buffer to store the Excel file
        output = io.BytesIO()

        # Import openpyxl styles for formatting
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils.dataframe import dataframe_to_rows
        from openpyxl.worksheet.table import Table, TableStyleInfo

        # Create Excel writer with formatting
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            workbook = writer.book

            # Define styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type="solid")
            currency_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                           top=Side(style='thin'), bottom=Side(style='thin'))

            # Export Investments with enhanced formatting
            investments = Investment.query.all()
            inv_data = [{
                'Date': i.date.strftime('%Y-%m-%d'),
                'Investor': i.investor_name,
                'Amount': i.amount,
                'Investment ID': f"INV-{i.id:03d}"
            } for i in investments]

            inv_df = pd.DataFrame(inv_data)
            inv_df.to_excel(writer, sheet_name='Investments', index=False)

            # Format Investment sheet
            inv_sheet = writer.sheets['Investments']
            inv_sheet.column_dimensions['A'].width = 12  # Date
            inv_sheet.column_dimensions['B'].width = 20  # Investor
            inv_sheet.column_dimensions['C'].width = 15  # Amount
            inv_sheet.column_dimensions['D'].width = 12  # Investment ID

            # Apply header formatting
            for col in ['A1', 'B1', 'C1', 'D1']:
                cell = inv_sheet[col]
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
                cell.border = border

            # Apply currency format and borders
            for row in range(2, len(inv_data) + 2):
                inv_sheet[f'C{row}'].number_format = '"£"#,##0.00'
                inv_sheet[f'C{row}'].fill = currency_fill
                for col in ['A', 'B', 'C', 'D']:
                    inv_sheet[f'{col}{row}'].border = border

            # Add table formatting
            if inv_data:
                inv_table = Table(displayName="InvestmentsTable", 
                                ref=f"A1:D{len(inv_data)+1}")
                inv_table.tableStyleInfo = TableStyleInfo(
                    name="TableStyleMedium9", showFirstColumn=False,
                    showLastColumn=False, showRowStripes=True, showColumnStripes=False)
                inv_sheet.add_table(inv_table)

            # Export Expenses with enhanced formatting
            expenses = Expense.query.all()
            exp_data = [{
                'Date': e.date.strftime('%Y-%m-%d'),
                'Description': e.description,
                'Category': e.category,
                'Amount': e.amount,
                'Expense ID': f"EXP-{e.id:03d}"
            } for e in expenses]

            exp_df = pd.DataFrame(exp_data)
            exp_df.to_excel(writer, sheet_name='Expenses', index=False)

            # Format Expense sheet
            exp_sheet = writer.sheets['Expenses']
            exp_sheet.column_dimensions['A'].width = 12  # Date
            exp_sheet.column_dimensions['B'].width = 35  # Description
            exp_sheet.column_dimensions['C'].width = 20  # Category
            exp_sheet.column_dimensions['D'].width = 15  # Amount
            exp_sheet.column_dimensions['E'].width = 12  # Expense ID

            # Apply header formatting
            for col in ['A1', 'B1', 'C1', 'D1', 'E1']:
                cell = exp_sheet[col]
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
                cell.border = border

            # Apply currency format and borders
            for row in range(2, len(exp_data) + 2):
                exp_sheet[f'D{row}'].number_format = '"£"#,##0.00'
                exp_sheet[f'D{row}'].fill = currency_fill
                for col in ['A', 'B', 'C', 'D', 'E']:
                    exp_sheet[f'{col}{row}'].border = border

            # Add table formatting
            if exp_data:
                exp_table = Table(displayName="ExpensesTable", 
                                ref=f"A1:E{len(exp_data)+1}")
                exp_table.tableStyleInfo = TableStyleInfo(
                    name="TableStyleMedium9", showFirstColumn=False,
                    showLastColumn=False, showRowStripes=True, showColumnStripes=False)
                exp_sheet.add_table(exp_table)

            # Export Sales with enhanced formatting
            sales = Sale.query.all()
            sales_data = [{
                'Date': s.date.strftime('%Y-%m-%d'),
                'Description': s.description,
                'Amount': s.amount,
                'Sale ID': f"SAL-{s.id:03d}"
            } for s in sales]

            sales_df = pd.DataFrame(sales_data)
            sales_df.to_excel(writer, sheet_name='Sales', index=False)

            # Format Sales sheet
            sales_sheet = writer.sheets['Sales']
            sales_sheet.column_dimensions['A'].width = 12  # Date
            sales_sheet.column_dimensions['B'].width = 35  # Description
            sales_sheet.column_dimensions['C'].width = 15  # Amount
            sales_sheet.column_dimensions['D'].width = 12  # Sale ID

            # Apply header formatting
            for col in ['A1', 'B1', 'C1', 'D1']:
                cell = sales_sheet[col]
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
                cell.border = border

            # Apply currency format and borders
            for row in range(2, len(sales_data) + 2):
                sales_sheet[f'C{row}'].number_format = '"£"#,##0.00'
                sales_sheet[f'C{row}'].fill = currency_fill
                for col in ['A', 'B', 'C', 'D']:
                    sales_sheet[f'{col}{row}'].border = border

            # Add table formatting
            if sales_data:
                sales_table = Table(displayName="SalesTable", 
                                ref=f"A1:D{len(sales_data)+1}")
                sales_table.tableStyleInfo = TableStyleInfo(
                    name="TableStyleMedium9", showFirstColumn=False,
                    showLastColumn=False, showRowStripes=True, showColumnStripes=False)
                sales_sheet.add_table(sales_table)

            # Create comprehensive Summary sheet
            total_investment = sum(i.amount for i in investments)
            total_expenses = sum(e.amount for e in expenses)
            total_sales = sum(s.amount for s in sales)
            net_profit = total_sales - total_expenses

            # Calculate investment shares
            shree_investment = sum(
                i.amount for i in investments if i.investor_name == 'Shree')
            adwait_investment = sum(
                i.amount for i in investments if i.investor_name == 'Adwait')

            # Calculate category breakdown for expenses
            expense_categories = {}
            for expense in expenses:
                if expense.category in expense_categories:
                    expense_categories[expense.category] += expense.amount
                else:
                    expense_categories[expense.category] = expense.amount

            # Create comprehensive summary data
            summary_data = [
                ['London\'s Kitchen - Complete Financial Report', ''],
                [f'Generated on: {datetime.now().strftime("%d/%m/%Y %H:%M")}', ''],
                ['', ''],
                ['📊 BUSINESS OVERVIEW', ''],
                ['Total Investments', total_investment],
                ['Total Expenses', total_expenses],
                ['Total Sales Revenue', total_sales],
                ['Net Profit/Loss', net_profit],
                ['Profit Margin', f"{(net_profit/total_sales*100):.1f}%" if total_sales > 0 else "0.0%"],
                ['', ''],
                ['💰 INVESTMENT ANALYSIS', ''],
                ['Total Number of Investments', len(investments)],
                ['Average Investment Amount', total_investment/len(investments) if investments else 0],
                ['Adwait\'s Total Investment', adwait_investment],
                ['Shree\'s Total Investment', shree_investment],
                ['Adwait\'s Investment %', f"{(adwait_investment/total_investment*100):.1f}%" if total_investment > 0 else "0.0%"],
                ['Shree\'s Investment %', f"{(shree_investment/total_investment*100):.1f}%" if total_investment > 0 else "0.0%"],
                ['', ''],
                ['💸 EXPENSE ANALYSIS', ''],
                ['Total Number of Expenses', len(expenses)],
                ['Average Expense Amount', total_expenses/len(expenses) if expenses else 0],
                ['', ''],
                ['Expense Categories Breakdown:', ''],
            ]

            # Add expense categories
            for category, amount in sorted(expense_categories.items(), key=lambda x: x[1], reverse=True):
                summary_data.append([f"  {category}", amount])

            # Continue with sales analysis
            summary_data.extend([
                ['', ''],
                ['💵 SALES ANALYSIS', ''],
                ['Total Number of Sales', len(sales)],
                ['Average Sale Amount', total_sales/len(sales) if sales else 0],
                ['', ''],
                ['📈 PROFIT DISTRIBUTION', ''],
                ['Available for Distribution', net_profit],
                ['Adwait\'s Profit Share', f"£{net_profit * (adwait_investment/total_investment):,.2f}" if total_investment > 0 else "£0.00"],
                ['Shree\'s Profit Share', f"£{net_profit * (shree_investment/total_investment):,.2f}" if total_investment > 0 else "£0.00"],
                ['', ''],
                ['🔍 DATA SUMMARY', ''],
                ['Total Records in System', len(investments) + len(expenses) + len(sales)],
                ['Data Export Date', datetime.now().strftime('%d/%m/%Y')],
                ['Data Export Time', datetime.now().strftime('%H:%M:%S')],
            ])

            summary_df = pd.DataFrame(summary_data, columns=['Category', 'Amount'])
            summary_df.to_excel(writer, sheet_name='Complete Summary', index=False)

            # Format Complete Summary sheet
            summary_sheet = writer.sheets['Complete Summary']
            summary_sheet.column_dimensions['A'].width = 40
            summary_sheet.column_dimensions['B'].width = 25

            # Apply enhanced formatting to Summary sheet
            summary_sheet['A1'].font = Font(bold=True, size=16, color="FFFFFF")
            summary_sheet['A1'].fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
            summary_sheet['A2'].font = Font(italic=True, size=10)

            # Apply section headers formatting
            section_headers = ['A4', 'A11', 'A20', 'A26', 'A34']
            for header in section_headers:
                cell = summary_sheet[header]
                cell.font = Font(bold=True, size=12, color="FFFFFF")
                cell.fill = PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type="solid")

            # Apply currency format and borders
            for row in range(1, len(summary_data) + 1):
                summary_sheet[f'A{row}'].border = border
                summary_sheet[f'B{row}'].border = border
                
                # Currency formatting for amount cells
                if row in [5, 6, 7, 8, 13, 14, 15, 16, 22, 23, 25, 26, 27, 28, 29, 35, 36, 37, 38]:
                    summary_sheet[f'B{row}'].number_format = '"£"#,##0.00'
                    summary_sheet[f'B{row}'].fill = currency_fill
                elif row in [9, 15, 16, 30]:
                    summary_sheet[f'B{row}'].number_format = '0.0%'

        output.seek(0)
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'London_Kitchen_Complete_Report_{datetime.now().strftime("%Y-%m-%d_%H-%M")}.xlsx'
        )
    except Exception as e:
        app.logger.error(f"Error in export_data: {str(e)}")
        flash(f"Error exporting data: {str(e)}", "error")
        return render_template('error.html', error=str(e))


@app.route('/dashboard')
def dashboard():
    total_investment = sum(inv.amount for inv in Investment.query.all())
    total_expenses = sum(exp.amount for exp in Expense.query.all())
    total_sales = sum(sale.amount for sale in Sale.query.all())
    net_profit_loss = total_sales - total_expenses

    # Calculate shares
    investments = Investment.query.all()
    shree_total = sum(
        i.amount for i in investments if i.investor_name == 'Shree')
    adwait_total = sum(
        i.amount for i in investments if i.investor_name == 'Adwait')

    # Get recent activity (last 10 items)
    recent_expenses = Expense.query.order_by(
        Expense.date.desc()).limit(5).all()
    recent_sales = Sale.query.order_by(Sale.date.desc()).limit(5).all()
    
    # Monthly sales analysis for the current year
    current_year = datetime.now().year
    monthly_sales = []
    monthly_expenses = []
    
    for month in range(1, 13):
        start_date = datetime(current_year, month, 1)
        if month == 12:
            end_date = datetime(current_year + 1, 1, 1) - timedelta(days=1)
        else:
            end_date = datetime(current_year, month + 1, 1) - timedelta(days=1)
            
        month_sales = Sale.query.filter(Sale.date >= start_date, Sale.date <= end_date).all()
        month_expenses = Expense.query.filter(Expense.date >= start_date, Expense.date <= end_date).all()
        
        monthly_sales.append({
            'month': calendar.month_name[month],
            'total': sum(sale.amount for sale in month_sales),
            'count': len(month_sales)
        })
        
        monthly_expenses.append({
            'month': calendar.month_name[month],
            'total': sum(expense.amount for expense in month_expenses),
            'count': len(month_expenses)
        })

    return render_template('dashboard.html',
                           total_investment=total_investment,
                           total_expenses=total_expenses,
                           total_sales=total_sales,
                           net_profit_loss=net_profit_loss,
                           adwait_investment=adwait_total,
                           shree_investment=shree_total,
                           recent_expenses=recent_expenses,
                           recent_sales=recent_sales,
                           monthly_sales=monthly_sales,
                           monthly_expenses=monthly_expenses,
                           current_year=current_year)


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)
